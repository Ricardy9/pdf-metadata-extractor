"""
PDF Metadata Extractor
Scans a folder of PDFs and writes a metadata spreadsheet to an .xlsx file.

Usage:
    python pdf_metadata_extractor.py /path/to/pdf/folder

Output:
    pdf_metadata.xlsx in the current working directory

Requirements:
    pip install pypdf openpyxl
"""

import multiprocessing
import os
import sys
import datetime
import re
from pathlib import Path

import pypdf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── Metadata extraction ──────────────────────────────────────────────────────

def _extract_metadata(pdf_path: Path, root_folder: Path) -> dict:
    """Extract metadata from a single PDF file."""
    result = {
        "filename":       pdf_path.name,
        "relative_path":  str(pdf_path.relative_to(root_folder)),
        "filepath":       str(pdf_path.resolve()),
        "file_size_kb":   round(pdf_path.stat().st_size / 1024, 1),
        "page_count":     None,
        "title":          None,
        "title_source":   None,
        "author":         None,
        "subject":        None,
        "creator":        None,
        "producer":       None,
        "creation_date":  None,
        "modified_date":  None,
        "word_count":     None,
        "has_text":       None,
        "error":          None,
    }

    try:
        # ── pypdf: page count + embedded metadata ────────────────────────────
        reader = pypdf.PdfReader(str(pdf_path))
        result["page_count"] = len(reader.pages)

        meta = reader.metadata or {}
        result["title"]    = _clean(meta.get("/Title"))
        result["author"]   = _clean(meta.get("/Author"))
        result["subject"]  = _clean(meta.get("/Subject"))
        result["creator"]  = _clean(meta.get("/Creator"))
        result["producer"] = _clean(meta.get("/Producer"))
        result["creation_date"] = _parse_pdf_date(meta.get("/CreationDate"))
        result["modified_date"] = _parse_pdf_date(meta.get("/ModDate"))

        # ── pypdf: text extraction + word count ───────────────────────────────
        all_text = ""
        for page in reader.pages:
            try:
                page_text = page.extract_text() or ""
            except Exception:
                page_text = ""
            all_text += page_text

        result["has_text"]  = len(all_text.strip()) > 0
        result["word_count"] = len(all_text.split()) if result["has_text"] else 0

        if result["title"]:
            result["title_source"] = "metadata"
        else:
            first_page_text = ""
            if reader.pages:
                try:
                    first_page_text = reader.pages[0].extract_text() or ""
                except Exception:
                    first_page_text = ""
            result["title"] = _extract_title_from_text(first_page_text)
            if result["title"]:
                result["title_source"] = "first_page_text"

    except Exception as exc:
        result["error"] = str(exc)

    return result


def _extract_metadata_worker(pdf_path_str: str, root_folder_str: str, queue: multiprocessing.Queue):
    pdf_path = Path(pdf_path_str)
    root_folder = Path(root_folder_str)
    try:
        result = _extract_metadata(pdf_path, root_folder)
    except Exception as exc:
        result = {
            "filename":       pdf_path.name,
            "relative_path":  str(pdf_path.relative_to(root_folder)) if root_folder in pdf_path.parents else str(pdf_path),
            "filepath":       str(pdf_path.resolve()),
            "file_size_kb":   round(pdf_path.stat().st_size / 1024, 1) if pdf_path.exists() else None,
            "page_count":     None,
            "title":          None,
            "title_source":   None,
            "author":         None,
            "subject":        None,
            "creator":         None,
            "producer":       None,
            "creation_date":  None,
            "modified_date": None,
            "word_count":     None,
            "has_text":       None,
            "error":          str(exc),
        }
    queue.put(result)


def extract_metadata(pdf_path: Path, root_folder: Path, timeout: int = 30) -> dict:
    queue = multiprocessing.Queue()
    process = multiprocessing.Process(
        target=_extract_metadata_worker,
        args=(str(pdf_path), str(root_folder), queue),
    )
    process.start()
    process.join(timeout)
    if process.is_alive():
        process.terminate()
        process.join()
        return {
            "filename":       pdf_path.name,
            "relative_path":  str(pdf_path.relative_to(root_folder)),
            "filepath":       str(pdf_path.resolve()),
            "file_size_kb":   round(pdf_path.stat().st_size / 1024, 1),
            "page_count":     None,
            "title":          None,
            "title_source":   None,
            "author":         None,
            "subject":        None,
            "creator":         None,
            "producer":       None,
            "creation_date":  None,
            "modified_date": None,
            "word_count":     None,
            "has_text":       None,
            "error":          f"timeout after {timeout} seconds",
        }

    if queue.empty():
        return {
            "filename":       pdf_path.name,
            "relative_path":  str(pdf_path.relative_to(root_folder)),
            "filepath":       str(pdf_path.resolve()),
            "file_size_kb":   round(pdf_path.stat().st_size / 1024, 1),
            "page_count":     None,
            "title":          None,
            "title_source":   None,
            "author":         None,
            "subject":        None,
            "creator":         None,
            "producer":       None,
            "creation_date":  None,
            "modified_date": None,
            "word_count":     None,
            "has_text":       None,
            "error":          "worker process failed without result",
        }

    return queue.get()


def _clean(value) -> str | None:
    """Strip whitespace from a metadata string; return None if empty."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _parse_pdf_date(raw) -> str | None:
    """
    Convert a PDF date string like "D:20230415120000+00'00'" to
    "2023-04-15 12:00:00", or return None if unparseable.
    """
    if not raw:
        return None
    s = str(raw).strip()
    if s.startswith("D:"):
        s = s[2:]
    # Extract numeric portion
    m = re.match(r"(\d{4})(\d{2})(\d{2})(\d{2})(\d{2})(\d{2})", s)
    if m:
        parts = m.groups()
        try:
            dt = datetime.datetime(*(int(p) for p in parts))
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass
    return s  # Return raw string if parsing fails


def _extract_title_from_text(text: str) -> str | None:
    """Return the first plausible title-like line from PDF text."""
    if not text:
        return None
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if line.isdigit():
            continue
        if re.fullmatch(r"page\s*\d+", line, re.IGNORECASE):
            continue
        if len(line) < 10:
            continue
        return line[:200]
    return None


# ── Spreadsheet creation ─────────────────────────────────────────────────────

COLUMNS = [
    ("filename",        "Filename",         28),
    ("relative_path",   "Relative Path",    40),
    ("page_count",      "Pages",             8),
    ("file_size_kb",    "File Size (KB)",   14),
    ("title",           "Title",            30),
    ("title_source",    "Title Source",     16),
    ("author",          "Author",           22),
    ("subject",         "Subject",          25),
    ("creator",         "Creator",          20),
    ("producer",        "Producer",         20),
    ("creation_date",   "Creation Date",    20),
    ("modified_date",   "Modified Date",    20),
    ("word_count",      "Word Count",       12),
    ("has_text",        "Has Text?",        11),
    ("filepath",        "Full Path",        50),
    ("error",           "Error",            35),
]

HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
ROW_FONT     = Font(name="Arial", size=10)
ALT_FILL     = PatternFill("solid", start_color="D6E4F0")
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=False)
ERROR_FONT   = Font(name="Arial", size=10, color="C00000")


def write_spreadsheet(records: list[dict], output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Metadata"

    # Header row
    for col_idx, (key, header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, rec in enumerate(records, start=2):
        is_alt = (row_idx % 2 == 0)
        for col_idx, (key, _header, _width) in enumerate(COLUMNS, start=1):
            value = rec.get(key)
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)

            if key == "error" and value:
                cell.font = ERROR_FONT
            else:
                cell.font = ROW_FONT

            cell.alignment = CENTER if key in ("page_count", "file_size_kb",
                                               "word_count", "has_text") else LEFT
            if is_alt and not (key == "error" and value):
                cell.fill = ALT_FILL

    # Summary row
    summary_row = len(records) + 2
    ws.cell(row=summary_row, column=1, value="TOTALS / AVERAGES").font = Font(
        name="Arial", bold=True, size=10)

    n = len(records)
    if n:
        pg_col  = get_column_letter(next(i for i,(k,*_) in enumerate(COLUMNS,1) if k=="page_count"))
        sz_col  = get_column_letter(next(i for i,(k,*_) in enumerate(COLUMNS,1) if k=="file_size_kb"))
        wc_col  = get_column_letter(next(i for i,(k,*_) in enumerate(COLUMNS,1) if k=="word_count"))
        data_start, data_end = 2, n + 1

        ws.cell(row=summary_row, column=COLUMNS.index(next(c for c in COLUMNS if c[0]=="page_count"))+1,
                value=f"=SUM({pg_col}{data_start}:{pg_col}{data_end})")
        ws.cell(row=summary_row, column=COLUMNS.index(next(c for c in COLUMNS if c[0]=="file_size_kb"))+1,
                value=f"=SUM({sz_col}{data_start}:{sz_col}{data_end})")
        ws.cell(row=summary_row, column=COLUMNS.index(next(c for c in COLUMNS if c[0]=="word_count"))+1,
                value=f"=SUM({wc_col}{data_start}:{wc_col}{data_end})")

    wb.save(str(output_path))
    print(f"✅  Saved: {output_path}  ({len(records)} PDFs)")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python pdf_metadata_extractor.py /path/to/pdf/folder")
        sys.exit(1)

    folder = Path(sys.argv[1])
    if not folder.is_dir():
        print(f"Error: '{folder}' is not a valid directory.")
        sys.exit(1)

    pdf_files = sorted(folder.glob("**/*.pdf"))  # recursive; use *.pdf for top-level only
    if not pdf_files:
        print(f"No PDF files found in: {folder}")
        sys.exit(0)

    print(f"Found {len(pdf_files)} PDF(s). Extracting metadata...")

    records = []
    for i, pdf_path in enumerate(pdf_files, start=1):
        print(f"  [{i}/{len(pdf_files)}] {pdf_path.name}", end="\r")
        records.append(extract_metadata(pdf_path, folder))

    print()  # newline after progress line

    output_path = Path("pdf_metadata.xlsx")
    write_spreadsheet(records, output_path)


if __name__ == "__main__":
    main()
